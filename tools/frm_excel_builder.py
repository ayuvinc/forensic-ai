"""FRM Excel builder (FR-04).

Produces a two-sheet .xlsx workbook from a list of RiskItem objects:
  Sheet 1 — Risk Register table (header + one row per risk item)
  Sheet 2 — Heat Map (5×5 grid, ARGB colour-coded by risk rating)

Writes atomically: saves to a .tmp file then os.replace() to final path.
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Sequence


# ARGB colours for the 5×5 heat map cells (likelihood × impact bands)
# Rating = likelihood × impact (1–25). Bands: 1-4 green, 5-9 amber, 10-25 red.
_GREEN  = "FF92D050"  # Excel-format ARGB — bright green
_AMBER  = "FFFFC000"  # amber
_RED    = "FFFF0000"  # red
_DARK_RED = "FFC00000"  # dark red for extreme (20–25)


def _rating_color(likelihood: int, impact: int) -> str:
    rating = likelihood * impact
    if rating >= 20:
        return _DARK_RED
    if rating >= 10:
        return _RED
    if rating >= 5:
        return _AMBER
    return _GREEN


class FRMExcelBuilder:
    """Builds a two-sheet FRM workbook from risk items.

    Usage:
        builder = FRMExcelBuilder()
        path = builder.build(risk_items, "cases/project/F_Final/risk_register.xlsx")
    """

    def build(self, risk_items: Sequence, output_path: str | Path) -> Path:
        """Write risk_items to a .xlsx workbook at output_path.

        AC: empty risk_items=[] produces a valid .xlsx with at least a header row.
        Returns the Path of the written file.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        # ── Sheet 1: Risk Register ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Risk Register"

        headers = [
            "Risk ID", "Category", "Title", "Description",
            "Likelihood", "Impact", "Rating", "Risk Owner",
            "Existing Controls", "Recommendations",
        ]
        header_fill = PatternFill(start_color="D50032", end_color="D50032", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)

        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Column widths
        widths = [10, 18, 25, 40, 12, 10, 10, 18, 30, 40]
        for col, w in enumerate(widths, 1):
            ws1.column_dimensions[ws1.cell(1, col).column_letter].width = w

        for row, item in enumerate(risk_items, 2):
            # Support both RiskItem objects and plain dicts
            def _get(attr: str, default=""):
                if hasattr(item, attr):
                    return getattr(item, attr)
                if isinstance(item, dict):
                    return item.get(attr, default)
                return default

            rating = _get("risk_rating") or (_get("likelihood", 1) * _get("impact", 1))
            rating_fill = PatternFill(
                start_color=_rating_color(_get("likelihood", 1), _get("impact", 1)).lstrip("FF"),
                end_color=_rating_color(_get("likelihood", 1), _get("impact", 1)).lstrip("FF"),
                fill_type="solid",
            )

            ws1.cell(row, 1, _get("risk_id"))
            ws1.cell(row, 2, _get("category"))
            ws1.cell(row, 3, _get("title"))
            ws1.cell(row, 4, _get("description"))
            ws1.cell(row, 5, _get("likelihood"))
            ws1.cell(row, 6, _get("impact"))
            rating_cell = ws1.cell(row, 7, rating)
            rating_cell.fill = rating_fill
            rating_cell.font = Font(bold=True)
            ws1.cell(row, 8, _get("risk_owner"))
            controls = _get("existing_controls", [])
            ws1.cell(row, 9, "; ".join(controls) if isinstance(controls, list) else str(controls))
            recs = _get("recommendations", [])
            ws1.cell(row, 10, "; ".join(recs) if isinstance(recs, list) else str(recs))

        # ── Sheet 2: Heat Map ─────────────────────────────────────────────────
        ws2 = wb.create_sheet("Heat Map")

        ws2.cell(1, 1, "Heat Map — Likelihood (rows) × Impact (columns)")
        ws2.cell(1, 1).font = Font(bold=True, size=11)
        ws2.merge_cells("A1:F1")

        # Column labels (impact 1–5)
        ws2.cell(2, 1, "L \\ I")
        for impact in range(1, 6):
            ws2.cell(2, impact + 1, f"Impact {impact}")
            ws2.cell(2, impact + 1).font = Font(bold=True)
            ws2.cell(2, impact + 1).alignment = Alignment(horizontal="center")

        # Row labels (likelihood 1–5) and coloured cells
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for likelihood in range(1, 6):
            row = likelihood + 2
            ws2.cell(row, 1, f"L={likelihood}")
            ws2.cell(row, 1).font = Font(bold=True)
            for impact in range(1, 6):
                col = impact + 1
                rating = likelihood * impact
                color = _rating_color(likelihood, impact)
                # openpyxl PatternFill takes 6-char hex (strips leading FF)
                hex6 = color[2:]
                cell = ws2.cell(row, col, rating)
                cell.fill      = PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")
                cell.font      = Font(bold=True, color="FFFFFF" if rating >= 5 else "000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = border

        for col in range(1, 7):
            ws2.column_dimensions[ws2.cell(2, col).column_letter].width = 12

        # Risk items plotted on map — mark the cell with item count
        counts: dict[tuple[int, int], int] = {}
        for item in risk_items:
            def _get(attr, default=1):
                if hasattr(item, attr):
                    return getattr(item, attr)
                return item.get(attr, default) if isinstance(item, dict) else default
            l = max(1, min(5, _get("likelihood")))
            i = max(1, min(5, _get("impact")))
            counts[(l, i)] = counts.get((l, i), 0) + 1

        for (l, i), count in counts.items():
            row = l + 2
            col = i + 1
            existing = ws2.cell(row, col).value or 0
            ws2.cell(row, col, f"{existing} ({count})")

        # ── Atomic write ───────────────────────────────────────────────────────
        tmp = out.with_suffix(".tmp.xlsx")
        wb.save(str(tmp))
        os.replace(tmp, out)
        return out
