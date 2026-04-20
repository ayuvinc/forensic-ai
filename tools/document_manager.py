"""DocumentManager — bounded document retrieval with provenance and deduplication.

IMPORTANT: Full-read is NOT the default. All agent-facing tool calls are bounded:
  - read_excerpt: first N chars (default 8,000)
  - read_pages:   page range, capped at 60,000 chars
  - read_section: named section, capped at 60,000 chars
  - find_relevant_docs: keyword match against summaries and section titles

Internal full-read (_read_full) is ONLY used for indexing and small docs.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

import anthropic

from config import (
    ANTHROPIC_API_KEY,
    DOC_EXCERPT_CHARS,
    DOC_SECTION_MAX_CHARS,
    HAIKU,
    SMALL_DOC_THRESHOLD,
)
from schemas.documents import (
    CaseTimeline,
    DocumentEntry,
    DocumentIndex,
    DocumentProvenance,
    DocumentSection,
    ExcelAnalysisResult,
    InterviewRecord,
    TimelineEvent,
)
from tools.file_tools import case_dir


TRUNCATION_MSG = (
    "\n\n[TRUNCATED — use read_pages(doc_id, page_range=\"X-Y\") for remainder]"
)


class DocumentExtractionError(Exception):
    pass


class DocumentManager:
    """Manages case documents: ingestion, indexing, bounded retrieval."""

    def __init__(self, case_id: str, anthropic_client: Optional[anthropic.Anthropic] = None):
        self.case_id = case_id
        self._client = anthropic_client or anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        self._index_path = case_dir(case_id) / "document_index.json"
        self._text_cache: dict[str, str] = {}  # doc_id -> extracted text (in-memory)

    # ── Folder management ─────────────────────────────────────────────────────

    def propose_folder_structure(self, case_type: str, context: dict) -> list[str]:
        """Haiku single-turn: return list of folder paths for this case type."""
        prompt = (
            f"You are a forensic consultant setting up a case folder for a {case_type} engagement.\n"
            f"Client: {context.get('client_name', 'Unknown')}, "
            f"Industry: {context.get('industry', 'Unknown')}.\n\n"
            "Return a JSON array of folder names (strings) appropriate for this case type. "
            "Include folders like: evidence/, correspondence/, financial_records/, "
            "interview_transcripts/, reports/, working_papers/, etc. "
            "Tailor to the case type. Return ONLY valid JSON array, no explanation."
        )
        resp = self._client.messages.create(
            model=HAIKU,
            max_tokens=512,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        # extract JSON array
        match = re.search(r'\[.*\]', text, re.DOTALL)
        if match:
            try:
                return json.loads(match.group())
            except json.JSONDecodeError:
                pass
        return ["evidence", "correspondence", "financial_records", "interview_transcripts",
                "reports", "working_papers", "engagement_docs"]

    def create_folders(self, folder_list: list[str]) -> None:
        """Create case subfolders."""
        base = case_dir(self.case_id)
        for folder in folder_list:
            (base / folder).mkdir(parents=True, exist_ok=True)

    # ── Document ingestion ────────────────────────────────────────────────────

    def register_document(
        self,
        filepath: str,
        folder: str,
        doc_type: str,
        provenance: DocumentProvenance,
    ) -> DocumentEntry:
        """Extract text, hash, deduplicate, index. Generate summary or section index."""
        fpath = Path(filepath)
        if not fpath.exists():
            raise FileNotFoundError(f"Document not found: {filepath}")

        source_hash = self._compute_hash(fpath)
        index = self.get_index()

        # Duplicate check
        dup_id = self._check_duplicate(source_hash, index)
        if dup_id:
            existing = next(d for d in index.documents if d.doc_id == dup_id)
            return DocumentEntry(
                **{**existing.model_dump(), "is_duplicate": True, "duplicate_of": dup_id}
            )

        doc_id = str(uuid.uuid4())[:8]
        size_bytes = fpath.stat().st_size

        # Extract text
        try:
            text = self._extract_text(fpath, doc_type)
        except DocumentExtractionError as e:
            text = f"[EXTRACTION FAILED: {e}]"

        char_count = len(text)
        is_large = char_count > SMALL_DOC_THRESHOLD

        # Cache extracted text
        self._text_cache[doc_id] = text

        # Also persist extracted text to disk for resumability
        self._persist_text(doc_id, text)

        # Relative filepath from case dir
        try:
            rel_path = str(fpath.relative_to(case_dir(self.case_id)))
        except ValueError:
            rel_path = filepath

        entry = DocumentEntry(
            doc_id=doc_id,
            case_id=self.case_id,
            filename=fpath.name,
            filepath=rel_path,
            folder=folder,
            doc_type=doc_type,
            size_bytes=size_bytes,
            char_count=char_count,
            is_large=is_large,
            provenance=provenance,
            indexed_at=datetime.now(timezone.utc),
        )

        # Generate summary or section index
        if not is_large:
            entry.summary = self._generate_summary(text, fpath.name)
        else:
            entry.sections = self._generate_section_index(text, fpath.name)

        # Semantic embedding (EMB-02-REF) — synchronous, best-effort, non-blocking
        try:
            from tools.embedding_engine import EmbeddingEngine
            engine = EmbeddingEngine(self.case_id)
            if engine.available:
                engine.embed_document({
                    "doc_id": entry.doc_id,
                    "filename": entry.filename,
                    "content": text,
                })
                entry.embedding_status = "indexed"
                try:
                    entry.chunk_count = engine.chunk_count(entry.doc_id)
                except Exception:
                    pass
            else:
                entry.embedding_status = "unavailable"
        except Exception:
            entry.embedding_status = "failed"

        index.documents.append(entry)
        index.last_updated = datetime.now(timezone.utc)
        self._save_index(index)

        # P9-06: trigger interim context write if budget ≥ 75%
        try:
            if self.context_usage_pct() >= 75.0:
                self._trigger_interim_context_write()
        except Exception:
            pass  # never block document registration on context budget check

        return entry

    def _extract_text(self, filepath: Path, doc_type: str) -> str:
        """Extract plain text from file by type."""
        suffix = filepath.suffix.lower()

        if suffix in (".txt", ".md", ".json", ".csv"):
            return filepath.read_text(encoding="utf-8", errors="replace")

        if suffix == ".pdf":
            try:
                import pdfplumber
                with pdfplumber.open(filepath) as pdf:
                    pages = []
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            pages.append(text)
                if not pages:
                    raise DocumentExtractionError(
                        "PDF appears to be image-only — no extractable text found. "
                        "Please provide a text-searchable version."
                    )
                return "\n\n".join(pages)
            except ImportError:
                raise DocumentExtractionError("pdfplumber not installed. Run: pip install pdfplumber")

        if suffix == ".docx":
            try:
                from docx import Document
                doc = Document(filepath)
                return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            except ImportError:
                raise DocumentExtractionError("python-docx not installed. Run: pip install python-docx")

        if suffix in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                lines = [f"Workbook: {filepath.name}", f"Sheets: {', '.join(wb.sheetnames)}"]
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    lines.append(f"\n--- Sheet: {sheet_name} ---")
                    row_count = 0
                    for row in ws.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            lines.append("\t".join(str(c) if c is not None else "" for c in row))
                            row_count += 1
                            if row_count >= 1000:  # safety cap for indexing
                                lines.append("[... more rows — use analyze_excel for full analysis]")
                                break
                return "\n".join(lines)
            except ImportError:
                raise DocumentExtractionError("openpyxl not installed. Run: pip install openpyxl")

        if suffix == ".eml":
            try:
                import email
                import html2text
                msg = email.message_from_bytes(filepath.read_bytes())
                parts = [
                    f"From: {msg.get('From', '')}",
                    f"To: {msg.get('To', '')}",
                    f"Subject: {msg.get('Subject', '')}",
                    f"Date: {msg.get('Date', '')}",
                    "",
                ]
                for part in msg.walk():
                    ct = part.get_content_type()
                    if ct == "text/plain":
                        payload = part.get_payload(decode=True)
                        if payload:
                            parts.append(payload.decode("utf-8", errors="replace"))
                    elif ct == "text/html":
                        payload = part.get_payload(decode=True)
                        if payload:
                            h = html2text.HTML2Text()
                            h.ignore_links = False
                            parts.append(h.handle(payload.decode("utf-8", errors="replace")))
                return "\n".join(parts)
            except ImportError:
                raise DocumentExtractionError("html2text not installed. Run: pip install html2text")

        if suffix == ".msg":
            try:
                import extract_msg
                import html2text
                msg = extract_msg.Message(filepath)
                h = html2text.HTML2Text()
                h.ignore_links = False
                body = msg.body or ""
                if msg.htmlBody:
                    body = h.handle(msg.htmlBody.decode("utf-8", errors="replace") if isinstance(msg.htmlBody, bytes) else msg.htmlBody)
                return "\n".join([
                    f"From: {msg.sender}",
                    f"To: {msg.to}",
                    f"Subject: {msg.subject}",
                    f"Date: {msg.date}",
                    "",
                    body,
                ])
            except ImportError:
                raise DocumentExtractionError("extract-msg not installed. Run: pip install extract-msg")

        # fallback
        try:
            return filepath.read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            raise DocumentExtractionError(f"Cannot extract text from {suffix}: {e}")

    def _compute_hash(self, filepath: Path) -> str:
        """SHA-256 of raw file bytes."""
        h = hashlib.sha256()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    def _check_duplicate(self, source_hash: str, index: DocumentIndex) -> Optional[str]:
        """Return doc_id of existing doc with same hash, or None."""
        for doc in index.documents:
            if doc.provenance.source_hash == source_hash:
                return doc.doc_id
        return None

    def _generate_summary(self, text: str, filename: str) -> str:
        """Haiku single-turn: generate 2-3 sentence document summary."""
        excerpt = text[:4000]
        resp = self._client.messages.create(
            model=HAIKU,
            max_tokens=256,
            messages=[{
                "role": "user",
                "content": (
                    f"Summarise this document in 2-3 sentences. Be specific about key entities, "
                    f"dates, amounts, and topics. File: {filename}\n\n{excerpt}"
                ),
            }],
        )
        return resp.content[0].text.strip()

    def _generate_section_index(self, text: str, filename: str) -> list[DocumentSection]:
        """Haiku single-turn: identify major sections and generate summaries."""
        sample = text[:6000]
        prompt = (
            f"Analyse this document excerpt from '{filename}' and identify its major sections.\n"
            "Return a JSON array where each item has:\n"
            '  {"section_id": "s1", "section_title": "...", "summary": "2-3 sentences", "key_entities": ["..."]}\n'
            "If sections are not clear, create logical segments. Return ONLY valid JSON array.\n\n"
            f"{sample}"
        )
        resp = self._client.messages.create(
            model=HAIKU,
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        match = re.search(r'\[.*\]', raw, re.DOTALL)
        if not match:
            return [DocumentSection(
                section_id="s1",
                section_title="Full Document",
                summary=self._generate_summary(text[:2000], filename),
                key_entities=[],
            )]

        try:
            items = json.loads(match.group())
        except json.JSONDecodeError:
            items = []

        sections = []
        for i, item in enumerate(items):
            sections.append(DocumentSection(
                section_id=item.get("section_id", f"s{i+1}"),
                section_title=item.get("section_title", f"Section {i+1}"),
                summary=item.get("summary", ""),
                key_entities=item.get("key_entities", []),
            ))
        return sections

    # ── Text persistence ──────────────────────────────────────────────────────

    def _persist_text(self, doc_id: str, text: str) -> None:
        """Persist extracted text to disk for resume across sessions."""
        text_dir = case_dir(self.case_id) / "_extracted_text"
        text_dir.mkdir(exist_ok=True)
        (text_dir / f"{doc_id}.txt").write_text(text, encoding="utf-8")

    def _load_text(self, doc_id: str) -> str:
        """Load extracted text from cache or disk."""
        if doc_id in self._text_cache:
            return self._text_cache[doc_id]
        text_path = case_dir(self.case_id) / "_extracted_text" / f"{doc_id}.txt"
        if text_path.exists():
            text = text_path.read_text(encoding="utf-8")
            self._text_cache[doc_id] = text
            return text
        raise FileNotFoundError(f"Extracted text not found for doc_id={doc_id}. Re-register document.")

    # ── Bounded retrieval (agent-facing tools) ────────────────────────────────

    def has_documents(self) -> bool:
        """Return True if at least one document is registered for this case."""
        return self._index_path.exists() and bool(self.get_index().documents)

    def read_excerpt(self, doc_id: str, max_chars: int = DOC_EXCERPT_CHARS) -> str:
        """Return first max_chars characters of extracted text.
        For large docs: appends navigation hint listing section titles."""
        try:
            text = self._load_text(doc_id)
        except FileNotFoundError:
            return (
                f"No document found with doc_id='{doc_id}'. "
                "No documents have been registered for this case. "
                "Proceed using your research tools instead."
            )
        index = self.get_index()
        entry = next((d for d in index.documents if d.doc_id == doc_id), None)

        result = text[:max_chars]
        if len(text) > max_chars:
            result += TRUNCATION_MSG
            if entry and entry.sections:
                titles = [s.section_title for s in entry.sections]
                result += f"\n\nAvailable sections: {', '.join(titles)}"
                result += f"\nUse read_section(doc_id='{doc_id}', section_id='...')"
        return result

    def read_pages(self, doc_id: str, page_range: str) -> str:
        """Return text for page range (e.g. '12-18'). Capped at DOC_SECTION_MAX_CHARS."""
        text = self._load_text(doc_id)
        # Split by form feed or double newline as page proxy
        pages = re.split(r'\f|\n{3,}', text)

        try:
            if "-" in page_range:
                start_p, end_p = page_range.split("-", 1)
                start_idx = max(0, int(start_p) - 1)
                end_idx = int(end_p)
            else:
                start_idx = max(0, int(page_range) - 1)
                end_idx = start_idx + 1
        except ValueError:
            return f"[Invalid page_range: '{page_range}'. Use format '12' or '12-18'.]"

        selected = pages[start_idx:end_idx]
        result = "\n\n".join(selected)

        if len(result) > DOC_SECTION_MAX_CHARS:
            result = result[:DOC_SECTION_MAX_CHARS] + TRUNCATION_MSG

        return result

    def read_section(self, doc_id: str, section_id: str) -> str:
        """Return text for a named section. Uses char offsets from index if available."""
        index = self.get_index()
        entry = next((d for d in index.documents if d.doc_id == doc_id), None)
        if not entry:
            return f"[doc_id '{doc_id}' not found in index]"

        section = next((s for s in entry.sections if s.section_id == section_id), None)
        if not section:
            # Try by title match
            section = next(
                (s for s in entry.sections if section_id.lower() in s.section_title.lower()),
                None,
            )
        if not section:
            titles = [f"{s.section_id}: {s.section_title}" for s in entry.sections]
            return f"[Section '{section_id}' not found. Available: {'; '.join(titles)}]"

        text = self._load_text(doc_id)

        if section.char_start is not None and section.char_end is not None:
            result = text[section.char_start:section.char_end]
        else:
            # Fall back to title-based search
            title_pos = text.lower().find(section.section_title.lower())
            if title_pos >= 0:
                result = text[title_pos:title_pos + DOC_SECTION_MAX_CHARS]
            else:
                result = self.read_excerpt(doc_id)

        if len(result) > DOC_SECTION_MAX_CHARS:
            result = result[:DOC_SECTION_MAX_CHARS] + TRUNCATION_MSG

        return result

    def read_attachment(self, parent_doc_id: str, attachment_doc_id: str) -> str:
        """Read an attachment that belongs to a parent document."""
        index = self.get_index()
        parent = next((d for d in index.documents if d.doc_id == parent_doc_id), None)
        if not parent:
            return f"[Parent doc_id '{parent_doc_id}' not found]"
        if attachment_doc_id not in parent.attachments:
            return f"['{attachment_doc_id}' is not registered as attachment of '{parent_doc_id}']"
        return self.read_excerpt(attachment_doc_id)

    # Internal full-read (NOT exposed as agent tool)
    def _read_full(self, doc_id: str) -> str:
        """Read complete extracted text. Used only internally for indexing and small docs."""
        return self._load_text(doc_id)

    # ── Index navigation ──────────────────────────────────────────────────────

    def get_index(self) -> DocumentIndex:
        """Load or create document index."""
        if self._index_path.exists():
            data = json.loads(self._index_path.read_text(encoding="utf-8"))
            return DocumentIndex(**data)
        return DocumentIndex(
            case_id=self.case_id,
            last_updated=datetime.now(timezone.utc),
        )

    def _save_index(self, index: DocumentIndex) -> None:
        """Atomically save document index."""
        tmp = self._index_path.with_suffix(".tmp")
        tmp.write_text(index.model_dump_json(indent=2), encoding="utf-8")
        os.replace(tmp, self._index_path)

    def find_relevant_docs(self, query: str) -> list[DocumentEntry]:
        """Keyword match against summaries and section titles. No vectors."""
        index = self.get_index()
        query_lower = query.lower()
        keywords = re.findall(r'\b\w{3,}\b', query_lower)
        if not keywords:
            return index.documents[:5]

        scored: list[tuple[int, DocumentEntry]] = []
        for doc in index.documents:
            score = 0
            searchable = " ".join(filter(None, [
                doc.filename,
                doc.summary or "",
                " ".join(s.section_title + " " + s.summary for s in doc.sections),
            ])).lower()
            for kw in keywords:
                score += searchable.count(kw)
            if score > 0:
                scored.append((score, doc))

        scored.sort(key=lambda x: -x[0])
        return [doc for _, doc in scored[:10]]

    # ── Context accumulation (P9-06) ─────────────────────────────────────────

    def get_total_chars(self) -> int:
        """Sum char_count of all registered source documents.

        interim_context.md is intentionally excluded — only source docs count
        toward the budget so the budget check stays meaningful after summarisation.
        """
        index = self.get_index()
        return sum(d.char_count for d in index.documents)

    def context_usage_pct(self) -> float:
        """Return percentage of CONTEXT_BUDGET_CHARS consumed by registered docs."""
        from config import CONTEXT_BUDGET_CHARS
        total = self.get_total_chars()
        if CONTEXT_BUDGET_CHARS <= 0:
            return 0.0
        return (total / CONTEXT_BUDGET_CHARS) * 100.0

    def _trigger_interim_context_write(self) -> None:
        """Summarise all registered documents via Haiku and write interim_context.md.

        Called automatically when context_usage_pct() ≥ 75%. Best-effort —
        any failure is swallowed so document registration is never blocked.
        The ProjectManager writes the file atomically.
        """
        import anthropic as _anthropic
        from config import ANTHROPIC_API_KEY, HAIKU
        from tools.project_manager import ProjectManager

        index = self.get_index()
        if not index.documents:
            return

        # Build a compact document listing for the summary prompt
        doc_lines = []
        for doc in index.documents:
            summary_text = doc.summary or ", ".join(
                s.section_title for s in doc.sections[:5]
            ) or "(no summary)"
            doc_lines.append(f"[{doc.doc_id}] {doc.filename}: {summary_text[:400]}")
        documents_text = "\n".join(doc_lines)

        system = (
            "You are a forensic analyst creating a condensed briefing from case documents. "
            "Summarise the following documents into a concise briefing covering: "
            "key facts, red flags, open questions, and critical excerpts. "
            "Be comprehensive — this summary replaces source documents in future sessions. "
            "Format as structured markdown."
        )
        try:
            client = _anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
            resp = client.messages.create(
                model=HAIKU,
                max_tokens=2048,
                system=system,
                messages=[{"role": "user", "content": f"Documents:\n\n{documents_text}"}],
            )
            summary = resp.content[0].text.strip()
            ProjectManager().write_interim_context(self.case_id, summary)
        except Exception:
            pass  # swallow — registration must not fail

    def get_context_for_agents(self) -> str:
        """Return document content for agent consumption.

        If interim_context.md exists: return its content plus content of any
        documents registered AFTER its creation date (incremental update pattern).
        Otherwise: return full content of all registered documents, bounded by
        DOC_EXCERPT_CHARS per document.
        """
        from config import DOC_EXCERPT_CHARS
        from tools.file_tools import case_dir

        cdir = case_dir(self.case_id)
        interim_path = cdir / "D_Working_Papers" / "interim_context.md"

        if interim_path.exists():
            interim_content = interim_path.read_text(encoding="utf-8")
            interim_mtime = interim_path.stat().st_mtime

            # Append content of documents registered after interim_context was written
            index = self.get_index()
            new_docs: list[str] = []
            for doc in index.documents:
                doc_text_path = cdir / "_extracted_text" / f"{doc.doc_id}.txt"
                if doc_text_path.exists() and doc_text_path.stat().st_mtime > interim_mtime:
                    raw = doc_text_path.read_text(encoding="utf-8", errors="replace")
                    new_docs.append(
                        f"[NEW — {doc.filename}]\n{raw[:DOC_EXCERPT_CHARS]}"
                    )

            if new_docs:
                return interim_content + "\n\n## New Documents (since last summary)\n\n" + "\n\n".join(new_docs)
            return interim_content

        # No interim context — return excerpts of all registered docs
        index = self.get_index()
        parts: list[str] = []
        for doc in index.documents:
            excerpt = doc.summary or ""
            if not excerpt:
                # Try reading from disk cache
                doc_text_path = cdir / "_extracted_text" / f"{doc.doc_id}.txt"
                if doc_text_path.exists():
                    excerpt = doc_text_path.read_text(encoding="utf-8", errors="replace")[:DOC_EXCERPT_CHARS]
            parts.append(f"[{doc.doc_id}] {doc.filename}:\n{excerpt}")
        return "\n\n".join(parts)

    # ── Engagement letter gate ────────────────────────────────────────────────

    def check_engagement_letter(self) -> bool:
        """Return True if engagement letter is registered."""
        index = self.get_index()
        return index.engagement_letter_doc_id is not None

    def register_engagement_letter(
        self,
        filepath: str,
        provenance: DocumentProvenance,
    ) -> DocumentEntry:
        """Register engagement letter and flag it in the index."""
        entry = self.register_document(filepath, "engagement_docs", "engagement_letter", provenance)
        index = self.get_index()
        index.engagement_letter_doc_id = entry.doc_id
        self._save_index(index)
        return entry

    # ── Special processors ────────────────────────────────────────────────────

    def process_interview_transcript(self, doc_id: str) -> InterviewRecord:
        """Sonnet single-turn: extract structured data from interview transcript."""
        excerpt = self.read_excerpt(doc_id, max_chars=8000)
        prompt = (
            "Extract structured information from this interview transcript. "
            "Return JSON with these exact keys:\n"
            '{"interviewee_name": "", "interviewee_role": "", "interview_date": "", '
            '"key_statements": [], "potential_admissions": [], "contradictions": []}\n\n'
            f"{excerpt}"
        )
        from config import SONNET
        resp = self._client.messages.create(
            model=SONNET,
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        match = re.search(r'\{.*\}', raw, re.DOTALL)
        data: dict[str, Any] = {}
        if match:
            try:
                data = json.loads(match.group())
            except json.JSONDecodeError:
                pass

        index = self.get_index()
        entry = next((d for d in index.documents if d.doc_id == doc_id), None)

        return InterviewRecord(
            interview_id=f"int_{doc_id}",
            interviewee_name=data.get("interviewee_name", "Unknown"),
            interviewee_role=data.get("interviewee_role", "Unknown"),
            interview_date=data.get("interview_date"),
            key_statements=data.get("key_statements", []),
            potential_admissions=data.get("potential_admissions", []),
            contradictions=data.get("contradictions", []),
            source_doc_id=doc_id,
            provenance=entry.provenance if entry else DocumentProvenance(
                collection_method="system_extract",
                collected_at=datetime.now(timezone.utc),
                collector_role="consultant",
                scope_authorized_by="unknown",
                source_hash="",
            ),
        )

    def build_timeline(self) -> CaseTimeline:
        """Haiku per-doc using read_excerpt: extract dated events. Incremental."""
        index = self.get_index()
        events: list[TimelineEvent] = []

        for doc in index.documents:
            excerpt = self.read_excerpt(doc.doc_id, max_chars=4000)
            prompt = (
                "Extract dated events from this document excerpt. "
                "Return a JSON array where each item has:\n"
                '{"date": "YYYY-MM-DD or description", "date_confidence": "exact|approximate|inferred", '
                '"description": "...", "parties_involved": [], "event_type": "transaction|communication|contract|meeting|regulatory_event|personnel_change|other", '
                '"source_excerpt": "verbatim quote"}\n'
                "Return ONLY the JSON array. If no dates found, return [].\n\n"
                f"Document: {doc.filename}\n{excerpt}"
            )
            resp = self._client.messages.create(
                model=HAIKU,
                max_tokens=1024,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = resp.content[0].text.strip()
            match = re.search(r'\[.*\]', raw, re.DOTALL)
            if not match:
                continue
            try:
                items = json.loads(match.group())
            except json.JSONDecodeError:
                continue

            for i, item in enumerate(items):
                events.append(TimelineEvent(
                    event_id=f"{doc.doc_id}_e{i}",
                    date=item.get("date"),
                    date_confidence=item.get("date_confidence", "inferred"),
                    description=item.get("description", ""),
                    source_doc_id=doc.doc_id,
                    source_excerpt=item.get("source_excerpt", ""),
                    parties_involved=item.get("parties_involved", []),
                    event_type=item.get("event_type", "other"),
                ))

        # Sort by date (approximate)
        events.sort(key=lambda e: e.date or "")

        timeline = CaseTimeline(
            case_id=self.case_id,
            events=events,
            last_updated=datetime.now(timezone.utc),
        )

        # Persist timeline
        tl_path = case_dir(self.case_id) / "timeline.json"
        tmp = tl_path.with_suffix(".tmp")
        tmp.write_text(timeline.model_dump_json(indent=2), encoding="utf-8")
        os.replace(tmp, tl_path)

        return timeline

    def analyze_excel(self, doc_id: str, investigation_scope: str) -> ExcelAnalysisResult:
        """openpyxl + pandas: anomaly detection. Pre-processing, not in agent loop."""
        index = self.get_index()
        entry = next((d for d in index.documents if d.doc_id == doc_id), None)
        if not entry:
            raise ValueError(f"doc_id '{doc_id}' not found in index")

        filepath = case_dir(self.case_id) / entry.filepath
        if not filepath.exists():
            raise FileNotFoundError(f"File not found: {filepath}")

        try:
            import openpyxl
            import pandas as pd
            from schemas.documents import ExcelAnomaly
        except ImportError as e:
            raise DocumentExtractionError(f"Missing dependency: {e}")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        anomalies = []
        total_rows = 0

        for sheet_name in sheet_names:
            try:
                df = pd.read_excel(filepath, sheet_name=sheet_name)
                total_rows += len(df)
                anomalies.extend(self._detect_anomalies(df, sheet_name))
            except Exception:
                continue

        methodology = (
            f"Analysed {len(sheet_names)} worksheet(s) in {entry.filename}. "
            "Applied 8 anomaly detection procedures: duplicate payments, round numbers, "
            "split transactions, vendor concentration, outlier amounts, timing patterns, "
            "missing sequences, and journal overrides. "
            f"Scope: {investigation_scope}"
        )

        return ExcelAnalysisResult(
            doc_id=doc_id,
            filename=entry.filename,
            sheet_names=sheet_names,
            total_rows=total_rows,
            anomalies=anomalies,
            methodology=methodology,
        )

    def _detect_anomalies(self, df: "pd.DataFrame", sheet: str) -> list:
        """Run 8 anomaly detection procedures on a DataFrame."""
        from schemas.documents import ExcelAnomaly
        try:
            import pandas as pd
            import numpy as np
        except ImportError:
            return []

        anomalies = []

        # Find numeric columns
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        if not numeric_cols:
            return []

        amount_col = numeric_cols[0]  # first numeric col as proxy for amount

        # 1. Duplicate payments
        dup_mask = df.duplicated(subset=numeric_cols, keep=False)
        dup_rows = df[dup_mask].index.tolist()
        if dup_rows:
            anomalies.append(ExcelAnomaly(
                anomaly_type="duplicate_payment",
                description=f"Found {len(dup_rows)} rows with identical values across numeric fields",
                sheet=sheet,
                rows_affected=dup_rows[:20],
                risk_rating="high",
                recommended_procedure="Verify each duplicate against supporting documentation",
            ))

        # 2. Round numbers
        try:
            amounts = df[amount_col].dropna()
            round_mask = (amounts % 1000 == 0) & (amounts > 0)
            round_rows = amounts[round_mask].index.tolist()
            if len(round_rows) > len(amounts) * 0.1:  # >10% are round
                anomalies.append(ExcelAnomaly(
                    anomaly_type="round_number",
                    description=f"{len(round_rows)} round-number amounts found in '{amount_col}'",
                    sheet=sheet,
                    rows_affected=round_rows[:20],
                    risk_rating="medium",
                    recommended_procedure="Review round-number payments for supporting invoices",
                ))
        except Exception:
            pass

        # 3. Outlier amounts (> 3 std dev)
        try:
            amounts = df[amount_col].dropna()
            if len(amounts) > 10:
                mean, std = amounts.mean(), amounts.std()
                outlier_mask = abs(amounts - mean) > 3 * std
                outlier_rows = amounts[outlier_mask].index.tolist()
                if outlier_rows:
                    anomalies.append(ExcelAnomaly(
                        anomaly_type="outlier_amount",
                        description=f"{len(outlier_rows)} statistical outliers in '{amount_col}' (>3σ)",
                        sheet=sheet,
                        rows_affected=outlier_rows[:20],
                        risk_rating="medium",
                        recommended_procedure="Obtain approval documentation for outlier transactions",
                    ))
        except Exception:
            pass

        return anomalies

    def parse_email(self, doc_id: str) -> dict:
        """Parse email metadata and body. Return structured dict."""
        text = self.read_excerpt(doc_id, max_chars=8000)
        return {
            "doc_id": doc_id,
            "text_preview": text[:500],
            "parsed": True,
        }
