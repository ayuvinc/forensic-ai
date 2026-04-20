"""Transaction Testing report section builders (WF-03).

TTSections builds structured markdown and Excel outputs for TT engagements.
build_excel_exceptions uses openpyxl (in requirements.txt) for .xlsx output.
"""
from __future__ import annotations

from pathlib import Path


class TTSections:
    """Builds structured output sections for Transaction Testing reports."""

    def build_exceptions_table(self, exceptions: list[dict]) -> str:
        """Return a markdown exceptions table.

        exceptions: list of dicts with keys transaction_id, date, amount,
                    description, exception_type, severity (all optional).
        """
        lines = ["## Exceptions Identified"]
        if not exceptions:
            lines.append("No exceptions identified in the testing population.")
            return "\n".join(lines)

        lines.append(
            "\n| # | Transaction ID | Date | Amount | Exception Type | Severity |"
        )
        lines.append("|---|---------------|------|--------|----------------|----------|")
        for i, exc in enumerate(exceptions, 1):
            tid      = exc.get("transaction_id", "—")
            date     = exc.get("date", "—")
            amount   = exc.get("amount", "—")
            etype    = exc.get("exception_type", "—")
            severity = exc.get("severity", "—").upper()
            lines.append(f"| {i} | {tid} | {date} | {amount} | {etype} | {severity} |")

        critical = sum(1 for e in exceptions if e.get("severity", "").lower() == "critical")
        high     = sum(1 for e in exceptions if e.get("severity", "").lower() == "high")
        lines.append(f"\n**Total exceptions:** {len(exceptions)}  |  Critical: {critical}  |  High: {high}")
        return "\n".join(lines)

    def build_summary_page(self, exceptions: list[dict], summary_data: dict) -> str:
        """Return a summary page suitable for embedding in a parent report.

        summary_data: dict with keys population_size, sample_size, testing_period,
                      scope_description (all optional).
        """
        lines = ["## Transaction Testing — Summary"]

        pop_size   = summary_data.get("population_size", "Not specified")
        sample     = summary_data.get("sample_size", "Not specified")
        period     = summary_data.get("testing_period", "Not specified")
        scope      = summary_data.get("scope_description", "")

        lines.append(f"\n**Testing Period:** {period}")
        lines.append(f"**Population Size:** {pop_size}")
        lines.append(f"**Sample Size:** {sample}")
        if scope:
            lines.append(f"\n{scope}")

        exception_rate = (
            round(len(exceptions) / int(sample) * 100, 1)
            if isinstance(sample, (int, float)) and sample > 0
            else "N/A"
        )
        lines.append(f"\n**Exceptions Found:** {len(exceptions)}")
        lines.append(f"**Exception Rate:** {exception_rate}%")

        if exceptions:
            lines.append("\n**Key exceptions (top 3):**")
            for exc in exceptions[:3]:
                tid   = exc.get("transaction_id", "—")
                etype = exc.get("exception_type", "—")
                lines.append(f"- {tid}: {etype}")

        return "\n".join(lines)

    def build_excel_exceptions(self, exceptions: list[dict], output_path: str | Path) -> Path:
        """Write exceptions to a .xlsx file with a header row.

        Sheet 1 contains: Transaction ID, Date, Amount, Description,
        Exception Type, Severity columns.

        AC: file is readable via openpyxl.load_workbook() without error.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exceptions"

        headers = ["Transaction ID", "Date", "Amount", "Description", "Exception Type", "Severity"]
        header_fill = PatternFill(start_color="D50032", end_color="D50032", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, exc in enumerate(exceptions, 2):
            ws.cell(row=row, column=1, value=exc.get("transaction_id", ""))
            ws.cell(row=row, column=2, value=exc.get("date", ""))
            ws.cell(row=row, column=3, value=exc.get("amount", ""))
            ws.cell(row=row, column=4, value=exc.get("description", ""))
            ws.cell(row=row, column=5, value=exc.get("exception_type", ""))
            ws.cell(row=row, column=6, value=exc.get("severity", ""))

        # Atomic write: save to .tmp, then rename
        tmp = out.with_suffix(".tmp.xlsx")
        wb.save(str(tmp))
        import os
        os.replace(tmp, out)
        return out
